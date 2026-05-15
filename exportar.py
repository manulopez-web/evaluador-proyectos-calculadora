"""
MÓDULO: EXPORTACIÓN AVANZADA

Este módulo permite exportar resultados financieros
a un archivo Excel profesional con:

- Tablas formateadas
- Indicadores financieros
- Gráficas automáticas
- Dashboard ejecutivo
- Análisis de sensibilidad
"""

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

# ---------------------------------------------------
# EXPORTAR EXCEL PROFESIONAL
# ---------------------------------------------------

def exportar_excel(df, sensibilidad_df=None):

    # ---------------------------------------------------
    # RUTA SEGURA (LOCAL + CLOUD)
    # ---------------------------------------------------

    carpeta = "exports"
    os.makedirs(carpeta, exist_ok=True)

    ruta = os.path.join(carpeta, "analisis_proyectos.xlsx")

    with pd.ExcelWriter(ruta, engine='openpyxl') as writer:

        # ---------------------------------------------
        # HOJA PRINCIPAL
        # ---------------------------------------------

        df.to_excel(
            writer,
            sheet_name="Resumen Ejecutivo",
            index=False
        )

        workbook = writer.book
        sheet = writer.sheets["Resumen Ejecutivo"]

        # ---------------------------------------------
        # ESTILOS
        # ---------------------------------------------

        color_titulo = PatternFill(
            start_color="0A2540",
            end_color="0A2540",
            fill_type="solid"
        )

        color_subtitulo = PatternFill(
            start_color="D9EAF7",
            end_color="D9EAF7",
            fill_type="solid"
        )

        fuente_blanca = Font(
            color="FFFFFF",
            bold=True,
            size=12
        )

        fuente_normal = Font(
            bold=True
        )

        # ---------------------------------------------
        # TÍTULO PRINCIPAL
        # ---------------------------------------------

        sheet.merge_cells("A1:H1")

        celda_titulo = sheet["A1"]

        celda_titulo.value = "REPORTE EJECUTIVO DE EVALUACIÓN DE PROYECTOS"

        celda_titulo.fill = color_titulo
        celda_titulo.font = Font(
            color="FFFFFF",
            bold=True,
            size=16
        )

        celda_titulo.alignment = Alignment(
            horizontal="center"
        )

        # ---------------------------------------------
        # FECHA
        # ---------------------------------------------

        sheet["A2"] = f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M')}"

        sheet["A2"].font = fuente_normal

        # ---------------------------------------------
        # FORMATO ENCABEZADOS
        # ---------------------------------------------

        # fila de encabezados (ajustada a pandas: fila 1 en Excel es encabezado real)
        header_row = 1

        for cell in sheet[header_row]:

            cell.fill = color_subtitulo
            cell.font = fuente_normal
            cell.alignment = Alignment(horizontal="center")

        # ---------------------------------------------
        # AJUSTAR COLUMNAS
        # ---------------------------------------------

        for col in sheet.columns:

            max_length = 0

            col_letter = get_column_letter(col[0].column)

            for cell in col:

                try:
                    if cell.value is not None:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))

                except:
                    pass

            adjusted_width = max_length + 5
            sheet.column_dimensions[col_letter].width = adjusted_width

        # ---------------------------------------------
        # GRÁFICA VPN
        # ---------------------------------------------

        grafica_vpn = BarChart()

        grafica_vpn.title = "Comparación de VPN"
        grafica_vpn.y_axis.title = "VPN"
        grafica_vpn.x_axis.title = "Proyectos"

        data = Reference(
            sheet,
            min_col=3,
            min_row=1,
            max_row=1 + len(df)
        )

        categorias = Reference(
            sheet,
            min_col=1,
            min_row=2,
            max_row=1 + len(df)
        )

        grafica_vpn.add_data(
            data,
            titles_from_data=True
        )

        grafica_vpn.set_categories(categorias)

        sheet.add_chart(
            grafica_vpn,
            "J5"
        )

        # ---------------------------------------------
        # GRÁFICA SCORE
        # ---------------------------------------------

        grafica_score = BarChart()

        grafica_score.title = "Comparación de Score"
        grafica_score.y_axis.title = "Score"
        grafica_score.x_axis.title = "Proyectos"

        data2 = Reference(
            sheet,
            min_col=7,
            min_row=1,
            max_row=1 + len(df)
        )

        grafica_score.add_data(
            data2,
            titles_from_data=True
        )

        grafica_score.set_categories(categorias)

        sheet.add_chart(
            grafica_score,
            "J20"
        )

        # ---------------------------------------------
        # SENSIBILIDAD
        # ---------------------------------------------

        if sensibilidad_df is not None:

            sensibilidad_df.to_excel(
                writer,
                sheet_name="Sensibilidad",
                index=False
            )

            sens_sheet = writer.sheets["Sensibilidad"]

            for cell in sens_sheet[1]:

                cell.fill = color_subtitulo
                cell.font = fuente_normal

            # Gráfica sensibilidad
            line_chart = LineChart()

            line_chart.title = "Sensibilidad VPN"
            line_chart.y_axis.title = "VPN"
            line_chart.x_axis.title = "Tasa (%)"

            data_line = Reference(
                sens_sheet,
                min_col=2,
                min_row=1,
                max_row=1 + len(sensibilidad_df)
            )

            cats_line = Reference(
                sens_sheet,
                min_col=1,
                min_row=2,
                max_row=1 + len(sensibilidad_df)
            )

            line_chart.add_data(
                data_line,
                titles_from_data=True
            )

            line_chart.set_categories(cats_line)

            sens_sheet.add_chart(
                line_chart,
                "E5"
            )

        # ---------------------------------------------
        # CONCLUSIÓN AUTOMÁTICA
        # ---------------------------------------------

        mejor = df.iloc[0]

        fila_final = len(df) + 8

        sheet[f"A{fila_final}"] = "Conclusión ejecutiva"

        sheet[f"A{fila_final}"].font = Font(
            bold=True,
            size=14
        )

        conclusion = f"""
El proyecto recomendado es {mejor['Proyecto']},
debido a que presenta el mayor score global,
combinando rentabilidad financiera e impacto social.
"""

        sheet[f"A{fila_final+1}"] = conclusion

        sheet[f"A{fila_final+1}"].alignment = Alignment(
            wrap_text=True
        )

    print("Archivo exportado correctamente.")
    print(f"Ubicación: {ruta}")

    return ruta